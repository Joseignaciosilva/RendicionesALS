from urllib.parse import urlparse
from django.conf import settings
from django.forms import ValidationError
from django.shortcuts import get_object_or_404, render
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework.decorators import api_view, action
from rest_framework import status
from .models import CentroCostos, Fondo, Gasto, Neteo
from .serializer import CentroCostosSerializer, FondoSerializer, GastoSerializer, NeteoSerializer
from rest_framework.parsers import MultiPartParser, FormParser
from django.core.files.storage import default_storage
from django.http import Http404, FileResponse, HttpResponse, JsonResponse
from rest_framework.decorators import api_view
from django.views.decorators.csrf import csrf_exempt
import os
import shutil
from django.core.files.storage import default_storage
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework import status
from rest_framework import viewsets
from .models import Fondo
from openpyxl import load_workbook  # type: ignore
from openpyxl.styles import Border, Side, Alignment, Font # type: ignore
from django.db.models import Q

    
class FondoViewSet(viewsets.ModelViewSet):
    queryset = Fondo.objects.all()
    serializer_class = FondoSerializer

    def get_queryset(self):
        # Obtener los parámetros de la consulta
        email = self.request.query_params.get('rendidor', None)
        rol = self.request.query_params.get('rol', None)
        usuario = self.request.query_params.get('usuario', None)  # Correo del usuario autenticado

        queryset = Fondo.objects.all()

        # Aplicar filtros según los parámetros presentes
        if email:
            queryset = queryset.filter(rendidor=email)
        
        if rol:
            if rol == 'JEF':
                queryset = queryset.filter(estado='en_jefatura', aprobadorJefatura=usuario)
            elif rol == 'ADM':
                queryset = queryset.filter(estado='en_administracion', aprobadorAdmin=usuario)

        return queryset



    
    @action(detail=True, methods=['delete'])
    def eliminar_fondo(self, request, pk=None):
        fondo = self.get_object()  # Obtener el fondo por su ID (pk)

        # Guardar información necesaria antes de eliminar el fondo
        fondo_id = fondo.id
        rendidor_email = fondo.rendidor.replace('@alsglobal.com', '').replace('.', '-')
        fondo_folder = os.path.join(settings.MEDIA_ROOT, 'imagenes', rendidor_email, str(fondo_id))
        rendidor_folder = os.path.join(settings.MEDIA_ROOT, 'imagenes', rendidor_email)

        # Eliminar los gastos y sus comprobantes
        for gasto in fondo.id_gastos.all():
            if gasto.nombreComprobante:
                try:
                    file_path = gasto.nombreComprobante.path
                    if os.path.exists(file_path):
                        default_storage.delete(file_path)  # Eliminar archivo
                        print(f"Comprobante {file_path} eliminado.")
                except Exception as e:
                    print(f"Error al eliminar la imagen del gasto: {e}")
            gasto.delete()  # Eliminar el gasto

        # Eliminar el fondo
        fondo.delete()
        # Eliminar la carpeta del fondo (id) y, si queda vacía, la carpeta del rendidor
        try:
            # Eliminar la carpeta del fondo
            if os.path.exists(fondo_folder):
                shutil.rmtree(fondo_folder)
                print(f"Carpeta {fondo_folder} eliminada correctamente.")
            # Eliminar carpeta del rendidor vacía
            if os.path.exists(rendidor_folder) and not os.listdir(rendidor_folder):
                shutil.rmtree(rendidor_folder)
                print(f"Carpeta {rendidor_folder} eliminada correctamente.")
        except Exception as e:
            print(f"Error al eliminar carpetas: {e}")

        return Response({"detail": "Fondo, gastos y carpetas eliminados correctamente."}, status=status.HTTP_204_NO_CONTENT)

    # Actualizar el total rendido del fondo
    @action(detail=True, methods=['post'])
    def actualizar_total_rendido(self, request, pk=None):
        try:
            fondo = self.get_object()
            total_rendido = request.data.get('totalRendido', None)

            if total_rendido is None:
                return Response({"error": "El total rendido es requerido."}, status=status.HTTP_400_BAD_REQUEST)

            fondo.totalRendido = total_rendido
            fondo.save()
            return Response({"message": "Total rendido actualizado con éxito."}, status=status.HTTP_200_OK)
        except Fondo.DoesNotExist:
            return Response({"error": "Fondo no encontrado."}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'])
    def calcular_neteo(self, request, pk=None):
        try:
            fondo = self.get_object()
            total_rendido = request.data.get('totalRendido', None)

            if total_rendido is None:
                return Response({"error": "El total rendido es requerido."}, status=status.HTTP_400_BAD_REQUEST)

            # Calcular saldos
            saldo_empresa = max(0, fondo.montoAsignado - total_rendido)
            saldo_rendidor = max(0, total_rendido - fondo.montoAsignado)

            # Actualizar o crear el registro de neteo
            neteo, created = Neteo.objects.update_or_create(
                idfondo=fondo,
                defaults={
                    "saldo_empresa": saldo_empresa,
                    "saldo_rendidor": saldo_rendidor
                }
            )

            return Response({
                "message": "Neteo calculado con éxito.",
                "saldo_empresa": saldo_empresa,
                "saldo_rendidor": saldo_rendidor
            }, status=status.HTTP_200_OK)

        except Fondo.DoesNotExist:
            return Response({"error": "Fondo no encontrado."}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'])
    def netear(self, request, pk=None):
        try:
            fondo = self.get_object()

            # Actualizar saldos del neteo
            neteo, created = Neteo.objects.update_or_create(
                idfondo=fondo,
                defaults={
                    "saldo_empresa": 0,
                    "saldo_rendidor": 0,
                }
            )
            if fondo.estado != 'en_cierre':
                return Response({"error": "El fondo no está en estado de cierre."}, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                "message": "Saldos neteados con éxito.",
                "saldo_empresa": neteo.saldo_empresa,
                "saldo_rendidor": neteo.saldo_rendidor
            }, status=status.HTTP_200_OK)
        except Fondo.DoesNotExist:
            return Response({"error": "Fondo no encontrado."}, status=status.HTTP_404_NOT_FOUND)

    def cerrar_fondo(fondo_id):
        try:
            fondo = Fondo.objects.get(id=fondo_id)

            # Verificar si los saldos están neteados
            if not hasattr(fondo, 'neteo'):
                raise ValidationError("El fondo no tiene un neteo asociado.")

            if fondo.neteo.saldo_empresa != 0 or fondo.neteo.saldo_rendidor != 0:
                raise ValidationError("El fondo no se puede cerrar hasta que los saldos estén neteados.")

            # Actualizar estado
            fondo.estado = "cerrado"
            fondo.save()
            return {"message": "Fondo cerrado con éxito."}
        except Fondo.DoesNotExist:
            raise ValidationError("El fondo especificado no existe.")
        except ValidationError as e:
            return {"error": str(e)}
 
class GastoViewSet(viewsets.ModelViewSet):
    queryset = Gasto.objects.all()
    serializer_class = GastoSerializer
    parser_classes = [MultiPartParser, FormParser]
    
    #suma del gasto al totalRendido
    def perform_create(self, serializer):
        gasto = serializer.save()
        fondo = gasto.idfondo
        fondo.update_total_rendido()
        return gasto
    
    #actualizar imagen, y totalRendido
    def perform_update(self, serializer):
        gasto = self.get_object()

        if 'nombreComprobante' in self.request.FILES:
            # Eliminar la imagen anterior si existe
            if gasto.nombreComprobante:
                try:
                    default_storage.delete(gasto.nombreComprobante.path)
                except Exception as e:
                    print(f"Error al eliminar la imagen anterior: {e}")
            gasto.nombreComprobante = self.request.FILES['nombreComprobante']  # Asignar la nueva imagen

        gasto = serializer.save() 
        fondo = gasto.idfondo 
        fondo.update_total_rendido() 

        return gasto
    
    # Eliminar imagen, totalRendido y el gasto
    def perform_destroy(self, instance):
        fondo = instance.idfondo
        if instance.nombreComprobante: # Eliminar la imagen asociada al gasto
            try:
                default_storage.delete(instance.nombreComprobante.path)
            except Exception as e:
                print(f"Error al eliminar la imagen: {e}")
        instance.delete() # Eliminar el gasto
        fondo.update_total_rendido() # Actualizar el total rendido del fondo
    
    # Crear el gasto
    def post(self, request, *args, **kwargs):
        serializer = GastoSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    #Accion personalizada de Get de idfondo para los gastos
    @action(detail=False, methods=['get'])
    def getFondo(self, request):
        fondo_id = request.query_params.get('idfondo')
        if fondo_id is not None:
            gastos = Gasto.objects.filter(idfondo=fondo_id)
            serializer = GastoSerializer(gastos, many=True)
            return Response(serializer.data)
        else:
            return Response({"detail": "Fondo ID no proporcionado."}, status=400)
    
    # Aprobar/rechazar gastos (acciones de jefe/admin)
    @action(detail=True, methods=['post'])
    def aprobar_jefe(self, request, pk=None):
        return self._cambiar_estado(request, pk, 'visadoJefe', 'aprobado')

    @action(detail=True, methods=['post'])
    def rechazar_jefe(self, request, pk=None):
        return self._cambiar_estado(request, pk, 'visadoJefe', 'rechazado')

    @action(detail=True, methods=['post'])
    def aprobar_admin(self, request, pk=None):
        return self._cambiar_estado(request, pk, 'visadoAdmin', 'aprobado')

    @action(detail=True, methods=['post'])
    def rechazar_admin(self, request, pk=None):
        return self._cambiar_estado(request, pk, 'visadoAdmin', 'rechazado')

    # Método auxiliar para cambiar estado de (visadoJefe / visadoAdmin)
    def _cambiar_estado(self, request, pk, campo, estado):
        try:
            gasto = self.get_object()
            setattr(gasto, campo, estado)
            gasto.save()
            return Response({"message": f"Gasto marcado como {estado} con éxito."}, status=status.HTTP_200_OK)
        except Gasto.DoesNotExist:
            return Response({"error": "Gasto no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        

class NeteoViewSet(viewsets.ModelViewSet):
    queryset = Neteo.objects.all()
    serializer_class = NeteoSerializer
    

class CentroCostosViewSet(viewsets.ModelViewSet):
    queryset = CentroCostos.objects.all()
    serializer_class = CentroCostosSerializer
        
    
@api_view(['POST'])
def getfactura(request):
    tipo_comprobante = request.data.get('tipoComprobante', '') 
    gastos = Gasto.objects.filter(
        Q(idfondo__estado='en_cierre') | Q(idfondo__estado='finalizado'),  
        tipoComprobante=tipo_comprobante
    ).select_related('idfondo')  
    if not gastos.exists():
        return Response(
        {"message": "No hay gastos con ese tipoComprobante en fondos 'en_cierre' o 'finalizado'."},
        status=status.HTTP_404_NOT_FOUND
    )
    serializer_gasto = GastoSerializer(gastos, many=True)
    return Response(serializer_gasto.data, status=status.HTTP_200_OK)


def descargar_comprobante(request, nombre_comprobante):
    # Construir la ruta al archivo dentro del directorio 'media/imagenes/'
    comprobante_path = os.path.join(settings.MEDIA_ROOT, 'imagenes', nombre_comprobante)

    # Verificar si el archivo existe
    if not os.path.exists(comprobante_path):
        return HttpResponse("Archivo no encontrado", status=404)

    # Devolver el archivo como respuesta para su descarga
    return FileResponse(open(comprobante_path, 'rb'), as_attachment=True, filename=nombre_comprobante)


def export_detalle_rendiciones(request, fondo_id):
    
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    plantilla_path = os.path.join(BASE_DIR, "api_rendiciones", "excel", "detalle_rendiciones.xlsx")
    
    wb = load_workbook(plantilla_path)
    ws = wb.active
    
    # Bordes y alineación
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Obtener los datos del fondo y gastos
    fondo = Fondo.objects.get(id=fondo_id)
    gastos = Gasto.objects.filter(idfondo=fondo)
    
    # Rellenar la plantilla con los datos del fondo
    ws['F17'] = fondo.rendidor
    ws['P17'] = fondo.rut
    ws['F19'] = fondo.aprobadorJefatura
    ws['P19'] = fondo.aprobadorAdmin
    ws['F25'] = fondo.referencia
    ws['F27'] = fondo.centroCosto or "N/A"
    ws['F29'] = fondo.montoAsignado
    ws['P25'] = fondo.fechaAsignado
    ws['P27'] = fondo.motoNave
    ws['P29'] = fondo.totalRendido
    
    # Fila inicial donde comienza la tabla de gastos
    fila_inicial = 37  
    altura_fila = ws.row_dimensions[fila_inicial].height  # Tomamos la altura de la primera fila de la tabla

    # Extraer tamaño de letra y alineación de cada celda en la fila base
    formatos_base = {}
    columnas = ["A", "C", "E", "G", "J", "L", "N", "P", "R", "T"]  # Columnas clave

    for col in columnas:
        celda_base = ws[f"{col}{fila_inicial}"]
        formatos_base[col] = {
            "fuente": celda_base.font.copy(),  # Crear copia para evitar el error
            "alineacion": celda_base.alignment.copy()
        }

    
    # Rellenar la tabla de gastos manteniendo el formato
    for i, gasto in enumerate(gastos, start=fila_inicial):
        ws.row_dimensions[i].height = altura_fila  # Copiar altura de la fila base
        
        campos = [
            ("A", "B", gasto.numeroServicio or "N/A"),
            ("C", "D", gasto.numeroComprobante),
            ("E", "F", gasto.tipoComprobante),
            ("G", "I", gasto.descripcion),
            ("J", "K", gasto.proveedor),
            ("L", "M", gasto.tipoGasto),
            ("N", "O", gasto.fechaGasto.strftime('%Y-%m-%d')),
            ("P", "Q", gasto.visadoJefe or "N/A"),
            ("R", "S", gasto.visadoAdmin or "N/A"),
            ("T", "U", gasto.montoGasto),
        ]

        for col_start, col_end, value in campos:
            celda_inicio = f"{col_start}{i}"
            ws[celda_inicio] = value
            
            # Aplicar formato de la plantilla
            if col_start in formatos_base:
                ws[celda_inicio].font = formatos_base[col_start]["fuente"]  # Copiar tamaño de letra original
                ws[celda_inicio].alignment = formatos_base[col_start]["alineacion"]
            
            # Asegurar alineación y ajuste de texto
            ws[celda_inicio].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Fusionar celdas
            ws.merge_cells(f"{col_start}{i}:{col_end}{i}")
            
            # Aplicar bordes a todas las celdas fusionadas
            for col in range(ws[col_start + str(i)].column, ws[col_end + str(i)].column + 1):
                ws.cell(row=i, column=col).border = thin_border

    # Generar respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="detalle_rendiciones_{fondo_id}.xlsx"'
    wb.save(response)
    
    return response






